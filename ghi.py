#!/usr/bin/python3
import collections
import configparser
import datetime
from enum import Enum
from functools import total_ordering
import glob
from http.server import HTTPServer, BaseHTTPRequestHandler
import itertools
import os
from threading import Thread
import tkinter
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from urllib.parse import urlparse
import webbrowser

import openpyxl
from requests_oauthlib import OAuth2Session


MONTH_NAME_BY_NUMBER = {1: 'Januar', 2: 'Februar', 3: 'March', 4: 'April', 5: 'Mai', 6: 'Juni', 7: 'Juli', 8: 'August', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Dezember'}
MONTH_NUMBER_BY_LOWER_NAME = {v.lower(): k for k, v in MONTH_NAME_BY_NUMBER.items()}
DAY_SHIFT_START = datetime.time(10)
NIGHT_SHIFT_START = datetime.time(20)
SHIFT_TIME_STRINGS = frozenset(t.strftime('%H:%M') for t in (DAY_SHIFT_START, NIGHT_SHIFT_START))


@total_ordering
class Shift(Enum):
	day = 0
	night = 1

	def __lt__(self, other):
		if self.__class__ is other.__class__:
			return self.value < other.value
		return NotImplemented


Midwife = collections.namedtuple('Midwife', ['name', 'phone'])
Duty = collections.namedtuple('Duty', ['date', 'shift'])
AttendantRule = collections.namedtuple('AttendantRule', ['id', 'schedule_name', 'schedule_id', 'event_by_duty'])


def ShiftName(shift):
	return {Shift.day: 'Tag', Shift.night: 'Nacht'}[shift]


def DutyName(duty):
	return f'{duty.date:%Y-%m-%d}{ShiftName(duty.shift)[0]}'


def DutyBeginEnd(duty):
	if duty.shift == Shift.day:
		return datetime.datetime.combine(duty.date, DAY_SHIFT_START), datetime.datetime.combine(duty.date, NIGHT_SHIFT_START)
	return datetime.datetime.combine(duty.date, NIGHT_SHIFT_START), datetime.datetime.combine(duty.date+datetime.timedelta(days=1), DAY_SHIFT_START)


def _ReadString(cell_value):
	if cell_value is not None:
		return str(cell_value).strip()
	return ''


class DutyRoster:
	def __init__(self, dates):
		self.dates = dates
		self.midwife_by_duty = {}
		self.midwife_by_name = {}

	def AddMidwife(self, midwife):
		name = midwife.name.lower()
		if name in self.midwife_by_name:
			raise RuntimeError(f'Hebamme {midwife} wurde schon angelegt!')
		self.midwife_by_name[name] = midwife

	def Add(self, midwife, duty):
		if duty in self.midwife_by_duty:
			raise RuntimeError(f'Dienst {duty} wurde schon angelegt!')
		self.midwife_by_duty[duty] = midwife

	def Check(self):
		duties = {Duty(d, st) for d in self.dates for st in Shift}
		duties.difference_update(self.midwife_by_duty)
		return [f'Niemand hat {ShiftName(d.shift)}schicht am {d.date:%d.%m.%Y}' for d in sorted(duties)]

	def GetMidwife(self, name):
		return self.midwife_by_name.get(name.lower())

	@staticmethod
	def FromWorkbook(wb_fn, sheet_index, start_row, start_col, primary_duty_tags):
		wb = openpyxl.load_workbook(wb_fn, data_only=True)
		sheet = wb.worksheets[sheet_index]
		month, year = sheet.oddHeader.left.text.strip().split()
		beginning_of_month = datetime.date(int(year), MONTH_NUMBER_BY_LOWER_NAME[month.lower()], 1)
		day = int(sheet.cell(start_row, start_col+1).value)
		if day > 20:
			dates = [(beginning_of_month-datetime.timedelta(days=1)).replace(day=day)]
		else:
			dates = [beginning_of_month.replace(day=day)]
		for end_col in range(start_col+1, start_col+100):
			try:
				day = int(sheet.cell(start_row, end_col+1).value)
			except Exception:
				end_col -= 1
				break
			dates.append(dates[-1]+datetime.timedelta(days=1))
			if day != dates[-1].day:
				raise RuntimeError(f'Ungültige Tage in Dienstplan: in Spalte {end_col} wurde {dates[-1].day} erwartet, aber {day} gefunden!')
		else:
			raise RuntimeError('Zu viele Spalten in Dienstplan!')
		duty_roster = DutyRoster(dates)
		for row in range(start_row+2, start_row+100, 2):
			name = _ReadString(sheet.cell(row, start_col).value)
			if not name:
				break
			midwife = Midwife(name, _ReadString(sheet.cell(row+1, start_col).value).replace(' ', '').replace('-', ''))
			try:
				int(midwife.phone)
			except ValueError:
				break
			duty_roster.AddMidwife(midwife)
			for j, date in enumerate(duty_roster.dates):
				for offset, st in [(0, Shift.day), (1, Shift.night)]:
					if tag := _ReadString(sheet.cell(row+offset, start_col+1+j).value).lower():
						if tag in primary_duty_tags:
							duty_roster.Add(midwife, Duty(date, st))
		return duty_roster


def NormalisePath(path):
	path = os.path.expanduser(path)
	if os.path.isabs(path):
		return path
	return os.path.normpath(os.path.join(os.path.dirname(__file__), path))


class Config:
	def __init__(self, config_fn):
		cp = configparser.ConfigParser()
		cp.read(config_fn, encoding='utf8')
		wb_cfg = cp['Workbook']
		self.sheet_index = wb_cfg.getint('sheet_index')
		self.start_row = wb_cfg.getint('start_row')
		self.start_col = wb_cfg.getint('start_col')
		self.primary_duty_tags = frozenset(t.strip().lower() for t in wb_cfg['primary_duty_tags'].split(',') if t.strip())
		self.wb_fn_pattern = wb_cfg['filename_pattern']
		self.day_shift_start_time = datetime.datetime.strptime(cp.get('Shifts', 'day_start'), '%H:%M')
		self.day_shift_end_time = datetime.datetime.strptime(cp.get('Shifts', 'day_end'), '%H:%M')
		gui_cfg = cp['GUI']
		self.table_col_width = gui_cfg.getint('table_col_width')
		self.n_midwifes_per_row = gui_cfg.getint('n_midwifes_per_row')
		self.new_forwarding_color = gui_cfg['new_forwarding_color']
		self.duty_forwarding_mismatch_color = gui_cfg['duty_forwarding_mismatch_color']
		self.webex_integration = cp['WebexIntegration']


class AutoAttendant:
	def __init__(self, attendant_id, location_id):
		self.id = attendant_id
		self.location_id = location_id
		self.rule_by_midwife = {}

	def AddRule(self, midwife, attendant_rule):
		if midwife in self.rule_by_midwife:
			raise RuntimeError(f'Mehrere Regeln für {midwife} gefunden!')
		self.rule_by_midwife[midwife] = attendant_rule

	def ForwardingRoster(self):
		dates = set()
		for ar in self.rule_by_midwife.values():
			dates.update(d.date for d in ar.event_by_duty)
		roster = DutyRoster(sorted(dates))
		for midwife, attendant_rule in self.rule_by_midwife.items():
			roster.AddMidwife(midwife)
			for duty in attendant_rule.event_by_duty:
				roster.Add(midwife, duty)
		return roster


class _Handler(BaseHTTPRequestHandler):
	def do_GET(self):
		self.server.last_request_path = self.path
		self.send_response(200)
		self.send_header('Content-type', 'text/html; charset=utf-8')
		self.end_headers()
		self.wfile.write('Das Code für das Token wurde erfolgreich abgerufen.'.encode())


class WebexApi:
	scopes = ('spark-admin:telephony_config_read', 'spark-admin:telephony_config_write')
	base_url = 'https://webexapis.com/v1/telephony/config'

	def __init__(self, integration_config, token_fn):
		self._client_id = integration_config['client_id']
		self._secret = integration_config['client_secret']
		self.redirect_uri = integration_config['redirect_uri']
		self.token_fn = token_fn
		self.cfg = configparser.ConfigParser()
		self.cfg.read(token_fn, encoding='utf8')
		min_token_expiry = datetime.datetime.today()+datetime.timedelta(hours=12)
		try:
			if min_token_expiry < datetime.datetime.fromisoformat(self.cfg.get('Webex', 'refresh_token_expires_at')):
				token = {k: self.cfg.get('Webex', k) for k in ('access_token', 'token_type')}
			else:
				token = None
		except (configparser.NoSectionError, configparser.NoOptionError):
			token = None
		self.session = OAuth2Session(client_id=self._client_id, scope=self.scopes, redirect_uri=self.redirect_uri, token=token)
		if not token:
			self.GetAccessToken()
		elif token and datetime.datetime.fromisoformat(self.cfg.get('Webex', 'expires_at')) < min_token_expiry:
			self.RefreshToken()

	def _Token(self):
		access_token = self.cfg.get('Webex', 'access_token')
		if access_token:
			return {'access_token': access_token, self.cfg.get('Webex', 'token_type'): 'Bearer'}
		return None

	def SaveToken(self, token):
		now = datetime.datetime.today()
		if not self.cfg.has_section('Webex'):
			self.cfg.add_section('Webex')
		cfg = self.cfg['Webex']
		for key in ('access_token', 'token_type', 'refresh_token'):
			cfg[key] = token[key]
		for key in ('expires_in', 'refresh_token_expires_in'):
			cfg[key.replace('_in', '_at')] = (now+datetime.timedelta(seconds=token[key]-10)).strftime('%Y-%m-%d %H:%M')
		with open(self.token_fn, 'w', encoding='utf8') as fo:
			self.cfg.write(fo)

	def GetAccessToken(self):
		url, _ = self.session.authorization_url('https://webexapis.com/v1/authorize')
		messagebox.showinfo(GHI.title, 'Um die Zugangsberechtigung zu bestätigen, wird ein Browser-Fenster geöffnet.')
		server = HTTPServer(('localhost', urlparse(self.redirect_uri).port), _Handler)
		thread = Thread(target=server.handle_request)
		thread.start()
		webbrowser.open_new(url)
		thread.join()
		authorization_response = self.redirect_uri+server.last_request_path
		# authorization_response = input('Die vollständige Adresszeile aus dem Browser kopieren und hier einfügen: ')
		env_key = 'OAUTHLIB_INSECURE_TRANSPORT'
		env_value = os.environ.get(env_key)
		os.environ[env_key] = '1'
		try:
			token = self.session.fetch_token('https://webexapis.com/v1/access_token', authorization_response=authorization_response, client_secret=self._secret)
		finally:
			if env_value is None:
				del os.environ[env_key]
			else:
				os.environ[env_key] = env_value
		self.SaveToken(token)

	def RefreshToken(self):
		self.SaveToken(self.session.refresh_token('https://webexapis.com/v1/access_token', refresh_token=self.cfg.get('Webex', 'refresh_token'), client_id=self._client_id, client_secret=self._secret))

	def Get(self, url):
		response = self.session.get(f'{self.base_url}/{url}')
		response.raise_for_status()
		return response.json()

	def Delete(self, url):
		response = self.session.delete(f'{self.base_url}/{url}')
		response.raise_for_status()

	def Post(self, url, json):
		response = self.session.post(f'{self.base_url}/{url}', json=json)
		response.raise_for_status()
		return response.json()

	def GetDuties(self, location_id, schedule_id, min_duty_end_ts=None):
		event_by_duty = {}
		for event in self.Get(f'locations/{location_id}/schedules/holidays/{schedule_id}')['events']:
			end_ts = datetime.datetime.strptime(f'{event["endDate"]} {event["endTime"]}', '%Y-%m-%d %H:%M')
			if min_duty_end_ts and end_ts < min_duty_end_ts:
				self.Delete(f'locations/{location_id}/schedules/holidays/{schedule_id}/events/{event["id"]}')
				continue
			if event['startTime'] not in SHIFT_TIME_STRINGS:
				raise RuntimeError(f'Ungültige Startzeit {event["startTime"]} am {event["startDate"]}!')
			start_ts = datetime.datetime.strptime(f'{event["startDate"]} {event["startTime"]}', '%Y-%m-%d %H:%M')
			while start_ts < end_ts:
				if start_ts.time() == DAY_SHIFT_START:
					duty = Duty(start_ts.date(), Shift.day)
					start_ts = datetime.datetime.combine(duty.date, NIGHT_SHIFT_START)
				else:
					duty = Duty(start_ts.date(), Shift.night)
					start_ts = datetime.datetime.combine(duty.date+datetime.timedelta(days=1), DAY_SHIFT_START)
				if duty not in event_by_duty:
					event_by_duty[duty] = event
				else:
					raise RuntimeError(f'Es gibt mehrere Rufumleitungen für {duty}!')
				if start_ts > end_ts:
					raise RuntimeError(f'Ungültige Endzeit {event["endTime"]} am {event["startDate"]}!')
		return event_by_duty

	def GetAutoAttendant(self, min_duty_end_ts):
		jsons = self.Get('autoAttendants')['autoAttendants']
		if len(jsons) != 1:
			raise RuntimeError(f'Es muss genau einen Auto-Attendant geben. Gefunden: {jsons}')
		attendant = AutoAttendant(jsons[0]['id'], jsons[0]['locationId'])
		schedule_id_by_name = {}
		for schedule in self.Get(f'locations/{attendant.location_id}/schedules')['schedules']:
			if schedule['type'] == 'holidays':
				schedule_id_by_name[schedule['name']] = schedule['id']
		for rule in self.Get(f'locations/{attendant.location_id}/autoAttendants/{attendant.id}/callForwarding')['callForwarding']['rules']:
			if rule['enabled']:
				try:
					schedule_name = self.Get(f'locations/{attendant.location_id}/autoAttendants/{attendant.id}/callForwarding/selectiveRules/{rule["id"]}')['holidaySchedule']
					schedule_id = schedule_id_by_name[schedule_name]
					attendant.AddRule(Midwife(rule['name'], rule['forwardTo']), AttendantRule(rule['id'], schedule_name, schedule_id, self.GetDuties(attendant.location_id, schedule_id, min_duty_end_ts=min_duty_end_ts)))
				except Exception as ex:
					raise RuntimeError(f'Fehler für {rule["name"]}: {ex}')
		return attendant

	def Upload(self, auto_attendant, midwife_by_duty):
		remove_by_midwife = collections.defaultdict(list)
		new_duties_by_midwife = collections.defaultdict(list)
		for midwife, rule in auto_attendant.rule_by_midwife.items():
			if any(d in midwife_by_duty for d in rule.event_by_duty):
				duties_by_id = collections.defaultdict(list)
				for duty, event in rule.event_by_duty.items():
					duties_by_id[event['id']].append(duty)
				for event_id, ed in duties_by_id.items():
					if any(d in midwife_by_duty for d in ed):
						remove_by_midwife[midwife].append(event_id)
						keep = [d for d in ed if d not in midwife_by_duty]
						if keep:
							new_duties_by_midwife[midwife].extend(keep)
		for duty, mw in midwife_by_duty.items():
			new_duties_by_midwife[mw].append(duty)
		schedules_url = f'locations/{auto_attendant.location_id}/schedules/holidays'
		for midwife, event_ids in remove_by_midwife.items():
			rule = auto_attendant.rule_by_midwife[midwife]
			for event_id in event_ids:
				self.Delete(f'{schedules_url}/{rule.schedule_id}/events/{event_id}')
		for midwife, duties in new_duties_by_midwife.items():
			duties.sort()
			times = [DutyBeginEnd(d) for d in duties]
			merged_duties = []
			start = 0
			end = 1
			n_duties = len(duties)
			while end <= n_duties:
				while end < len(times) and times[end-1][1] == times[end][0]:
					end += 1
				merged_duties.append(duties[start:end])
				start = end
				end += 1
			events = []
			for event_duties in merged_duties:
				name = DutyName(event_duties[0])
				begin, end = DutyBeginEnd(event_duties[0])
				if len(event_duties) > 1:
					name = f'{name}-{DutyName(event_duties[-1])}'
					end = DutyBeginEnd(event_duties[-1])[1]
				events.append({'name': name, 'startDate': begin.strftime('%Y-%m-%d'), 'startTime': begin.strftime('%H:%M'), 'endDate': end.strftime('%Y-%m-%d'), 'endTime': end.strftime('%H:%M')})
			if midwife in auto_attendant.rule_by_midwife:
				schedule_id = auto_attendant.rule_by_midwife[midwife].schedule_id
				for event in events:
					self.Post(f'{schedules_url}/{schedule_id}/events', event)
			else:
				schedule_id = self.Post(f'locations/{auto_attendant.location_id}/schedules', {'type': 'holidays', 'name': midwife.name, 'events': events})['id']
				rule = {'name': midwife.name, 'enabled': True, 'holidaySchedule': midwife.name, 'forwardTo': {'phoneNumber': midwife.phone, 'selection': 'FORWARD_TO_SPECIFIED_NUMBER'}, 'callsFrom': {'selection': 'ANY'}}
				rule_id = self.Post(f'locations/{auto_attendant.location_id}/autoAttendants/{auto_attendant.id}/callForwarding/selectiveRules', rule)['id']
				auto_attendant.AddRule(midwife, AttendantRule(rule_id, midwife.name, schedule_id, {}))
		for rule in auto_attendant.rule_by_midwife.values():
			rule.event_by_duty.clear()
			rule.event_by_duty.update(self.GetDuties(auto_attendant.location_id, rule.schedule_id))


class BoxStyle:
	def __init__(self, box, forwarding):
		self.box = box
		self.forwarding = forwarding

	def OnSelected(self, _):
		self.box['style'] = 'TCombobox' if self.box.get() == self.forwarding else 'New.TCombobox'


class GHI:
	title = 'Geburtshaus Idstein - Rufbereitschaft'

	def __init__(self, config_fn, token_fn):
		self.config = Config(config_fn)
		self.api = WebexApi(self.config.webex_integration, token_fn)
		self.attendant = self.api.GetAutoAttendant(datetime.datetime.combine(datetime.date.today()-datetime.timedelta(days=2), NIGHT_SHIFT_START))
		self.forwarding_roster = None
		self.duty_roster = None
		self.box_by_duty = {}
		self.midwife_by_box_value = {}
		self.table_frame = None
		self.midwife_frame = None

	def StartGui(self):
		root = tkinter.Tk()
		style = ttk.Style()
		style.configure('New.TCombobox', foreground=self.config.new_forwarding_color, background=self.config.new_forwarding_color)
		root.title(self.title)
		self.table_frame = ttk.LabelFrame(root, text='Rufumleitungen')
		self.table_frame.grid(row=0, column=0, sticky='nswe', padx=2, pady=2)
		self.midwife_frame = ttk.LabelFrame(root, text='Hebammen')
		self.midwife_frame.grid(row=1, column=0, sticky='we', padx=2, pady=2)
		self.LoadForwardings()
		frame = ttk.Frame(root)
		ttk.Button(frame, text='Dienstplan laden', command=self.LoadWorkbook).grid(row=0, column=0, sticky='w')
		ttk.Button(frame, text='Rufumleitungen hochladen', command=self.UploadForwardings).grid(row=0, column=1, sticky='w')
		frame.grid(row=2, column=0, sticky='we')
		root.rowconfigure(0, weight=1)
		root.columnconfigure(0, weight=1)
		root.mainloop()

	def LoadForwardings(self):
		try:
			self.forwarding_roster = self.attendant.ForwardingRoster()
			self.UpdateWidgets()
		except Exception as ex:
			messagebox.showerror('Konnte Rufumleitungen nicht laden', str(ex), master=self.table_frame)

	def UploadForwardings(self):
		midwife_by_duty = {}
		for duty, box in self.box_by_duty.items():
			value = box.get()
			if value and self.forwarding_roster.midwife_by_duty.get(duty) != self.midwife_by_box_value[value]:
				midwife_by_duty[duty] = self.midwife_by_box_value[value]
		self.api.Upload(self.attendant, midwife_by_duty)
		self.LoadForwardings()

	def LoadWorkbook(self):
		try:
			wb_fn_pattern = os.path.expanduser(self.config.wb_fn_pattern)
			try:
				wb_fn = max(glob.iglob(wb_fn_pattern), key=os.path.getmtime)
			except ValueError:
				dlg_args = {'initialdir': os.path.dirname(wb_fn_pattern)}
			else:
				dlg_args = {'initialdir': os.path.dirname(wb_fn), 'initialfile': os.path.basename(wb_fn)}
			wb_fn = filedialog.askopenfilename(parent=self.table_frame, title='Dienstplan workbook', filetypes=[('Excel', '.xlsx')], **dlg_args)
			if not wb_fn:
				return
			duty_roster = DutyRoster.FromWorkbook(wb_fn, self.config.sheet_index, self.config.start_row, self.config.start_col, self.config.primary_duty_tags)
			self.duty_roster = duty_roster
			self.UpdateWidgets()
		except Exception as ex:
			messagebox.showerror('Konnte Dienstplan nicht laden', str(ex), master=self.table_frame)
			raise

	def UpdateWidgets(self):
		min_date = datetime.date.today()-datetime.timedelta(days=1)
		if self.duty_roster:
			dates = {d for d in self.duty_roster.dates if min_date <= d}
			midwifes = list(self.duty_roster.midwife_by_name.values())
		else:
			dates = set()
			midwifes = []
		dates.update(d for d in self.forwarding_roster.dates if min_date <= d)
		if midwifes:
			for midwife in self.forwarding_roster.midwife_by_name.values():
				mw = self.duty_roster.GetMidwife(midwife.name)
				if not mw:
					if midwife in self.forwarding_roster.midwife_by_duty.values():
						midwifes.append(midwife)
				elif mw != midwife:
					raise RuntimeError(f'Inkonsistente Hebammendaten: {midwife} und {mw}!')
		else:
			midwifes = list(self.forwarding_roster.midwife_by_name.values())
		if not dates:
			dates.add(min_date)
		dates = sorted(dates)
		midwifes.sort()
		self.midwife_by_box_value = {}
		for midwife in midwifes:
			items = midwife.name.split()
			for i in range(1, max(len(items[0]), len(items[-1]))):
				key = items[0][:i]+items[-1][:i]
				if key not in self.midwife_by_box_value:
					self.midwife_by_box_value[key] = midwife
					break
		key_by_midwife = {m: k for k, m in self.midwife_by_box_value.items()}
		self.box_by_duty.clear()
		for widget in itertools.chain(self.table_frame.winfo_children(), self.midwife_frame.winfo_children()):
			widget.destroy()
		col_by_date = {d: i+2 for i, d in enumerate(sorted(dates))}
		ttk.Label(self.table_frame, text=f'Start: {next(iter(dates)):%d.%m.%Y}', font='TkHeadingFont').grid(row=0, column=0)
		for date, col in col_by_date.items():
			ttk.Label(self.table_frame, text=str(date.day), font='TkHeadingFont', justify='center', width=self.config.table_col_width).grid(row=0, column=col)
		n_cols = 2+len(dates)
		ttk.Separator(self.table_frame, orient='horizontal').grid(row=1, column=0, sticky='we', columnspan=n_cols)
		box_row = 2
		if self.duty_roster and min_date <= self.duty_roster.dates[-1]:
			duty_row = box_row
			box_row += 3
		else:
			duty_row = None
		if self.forwarding_roster.dates and min_date <= self.forwarding_roster.dates[-1]:
			forwarding_row = box_row
			box_row += 3
		else:
			forwarding_row = None
		for label, row in [('Dienstplan', duty_row), ('Rufumleitung', forwarding_row)]:
			if row:
				ttk.Label(self.table_frame, text=label, font='TkHeadingFont').grid(row=row, column=0)
				ttk.Label(self.table_frame, text='T', font='TkHeadingFont').grid(row=row, column=1)
				ttk.Label(self.table_frame, text='N', font='TkHeadingFont').grid(row=row+1, column=1)
				ttk.Separator(self.table_frame, orient='horizontal').grid(row=row + 2, column=0, sticky='we', columnspan=n_cols)
		ttk.Label(self.table_frame, text='T', font='TkHeadingFont').grid(row=box_row, column=1)
		ttk.Label(self.table_frame, text='N', font='TkHeadingFont').grid(row=box_row+1, column=1)
		box_values = tuple(sorted(self.midwife_by_box_value))
		for date, col in col_by_date.items():
			for shift in Shift:
				duty = Duty(date, shift)
				if self.duty_roster and duty in self.duty_roster.midwife_by_duty:
					roster = key_by_midwife[self.duty_roster.midwife_by_duty[duty]]
				else:
					roster = None
				if self.forwarding_roster and duty in self.forwarding_roster.midwife_by_duty:
					forwarding = key_by_midwife[self.forwarding_roster.midwife_by_duty[duty]]
				else:
					forwarding = None
				color = {'background': self.config.duty_forwarding_mismatch_color} if roster != forwarding and roster and forwarding else {}
				for value, row in [(roster, duty_row), (forwarding, forwarding_row)]:
					if value:
						ttk.Label(self.table_frame, text=value, justify='center', width=self.config.table_col_width, **color).grid(row=row+duty.shift.value, column=col)
				self.box_by_duty[duty] = ttk.Combobox(self.table_frame, values=box_values, state='readonly', justify='center', width=self.config.table_col_width)
				value = forwarding or roster
				box = self.box_by_duty[duty]
				if value:
					box.set(value)
					if value != forwarding:
						box['style'] = 'New.TCombobox'
				box.bind('<<ComboboxSelected>>', BoxStyle(box, forwarding).OnSelected)
				box.grid(row=box_row+shift.value, column=col, sticky='we')
		self.table_frame.columnconfigure(tuple(range(1, n_cols+1)), weight=1)
		row = 0
		col = 0
		for key in box_values:
			ttk.Label(self.midwife_frame, text=f'{key}: {self.midwife_by_box_value[key].name}').grid(row=row, column=col, sticky='we')
			col += 1
			if col > self.config.n_midwifes_per_row:
				row += 1
				col = 0
		self.midwife_frame.columnconfigure(tuple(range(col if row == 0 else self.config.n_midwifes_per_row)), weight=1)


def Main():
	try:
		app_dir = os.path.dirname(__file__)
		GHI(os.path.join(app_dir, 'ghi.ini'), os.path.join(app_dir, 'token.ini')).StartGui()
	except Exception as ex:
		messagebox.showerror(GHI.title, f'Fehler beim Start: {ex.__class__.__name__}: {ex}')


if __name__ == '__main__':
	Main()
